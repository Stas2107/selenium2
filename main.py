from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook


def read_tags_from_file(filename):
    with open(filename, 'r', encoding='utf-8') as file:
        tags = [line.strip() for line in file.readlines()]
    return tags


def scrape_data(browser, tags):
    data = {tag: [] for tag in tags}

    for element in browser.find_elements(By.TAG_NAME, "div"):
        cl = element.get_attribute("class")
        for tag in tags:
            if cl and tag in cl:
                data[tag].append(element.text)
                break
    return data


# Настройка драйвера и переход на нужную страницу
browser = webdriver.Chrome()
browser.get('https://phdays.com/forum/program/?date=2024%2F5%2F23')
time.sleep(20)  # Даем время для полной загрузки страницы

# Чтение тегов из файла
tags = read_tags_from_file('tegs.txt')

# Скрапинг данных
data = scrape_data(browser, tags)

# Создаем новую книгу Excel
wb = Workbook()
ws = wb.active

# Записываем данные в столбцы Excel
for idx, (tag, values) in enumerate(data.items(), 1):
    ws[f'{chr(64 + idx)}1'] = tag  # Записываем имя тега в заголовок столбца
    for row_idx, value in enumerate(values, 2):  # Начинаем с 2-ой строки, так как 1-я строка под заголовки
        ws[f'{chr(64 + idx)}{row_idx}'] = value

# Сохраняем книгу
wb.save("doclades.xlsx")

# Закрываем браузер
browser.quit()