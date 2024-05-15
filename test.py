from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook

# Настройка драйвера и переход на нужную страницу
browser = webdriver.Chrome()
browser.get('https://phdays.com/forum/program/?date=2024%2F5%2F26')
time.sleep(10)  # Даем время для полной загрузки страницы

# Находим все элементы с нужным классом
doclades = []
talk_times = []
areas = []

for element in browser.find_elements(By.TAG_NAME, "div"):
    cl = element.get_attribute("class")
    if cl and "TalkCard_actions-area" in cl:
        areas.append(element.text)
    if cl and "TalkTime_talk-time" in cl:
        talk_times.append(element.text)
    if cl and "TalkInfo_details" in cl:
        doclades.append(element.text)

# Создаем новую книгу Excel
wb = Workbook()
ws = wb.active

# Записываем данные в первый, второй и третий столбцы
for idx, (area, talk_time, doclade) in enumerate(zip(areas, talk_times, doclades), 1):
    ws[f'A{idx}'] = area
    ws[f'B{idx}'] = talk_time
    ws[f'C{idx}'] = doclade

# Сохраняем книгу
wb.save("doclades.xlsx")

# Закрываем браузер
browser.quit()